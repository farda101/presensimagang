import atexit
import json
import os
import re
import uuid
from collections import Counter
from datetime import datetime, time, timedelta, timezone

import mysql.connector
import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
from flask import (Flask, jsonify, redirect, render_template, request,
                   send_file, session, url_for)
from openpyxl import Workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = '********'

# Database connection
def get_db_connection():
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="internship_diary_app"
    )
    conn.cursor(dictionary=True)
    return conn

# Scheduler function to update absent students to 'Alpha' after a certain time
def update_alpha_status():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    today = datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.now().time()
    batas_waktu_presensi = time(23, 59)  # Misalnya jam 9 pagi

    if current_time >= batas_waktu_presensi:
        # Ambil semua mahasiswa
        cursor.execute("SELECT id, nama FROM mahasiswa")
        mahasiswa_list = cursor.fetchall()

        # Ambil mahasiswa yang sudah presensi atau izin
        cursor.execute("""
            SELECT nama_mahasiswa FROM presensi_mahasiswa 
            WHERE DATE(waktu_presensi) = %s
            AND status IN ('Hadir', 'Sakit', 'Izin', 'Alpha')
        """, (today,))
        presensi_today = cursor.fetchall()

        # Daftar mahasiswa yang sudah presensi, izin, atau Alpha
        presensi_nama_mahasiswa = [presensi['nama_mahasiswa'] for presensi in presensi_today]

        # Insert mahasiswa yang belum presensi, izin, atau Alpha dengan status Alpha
        for mahasiswa in mahasiswa_list:
            if mahasiswa['nama'] not in presensi_nama_mahasiswa:
                waktu_presensi_alpha = datetime.combine(datetime.now().date(), batas_waktu_presensi)  # Tanggal hari ini + batas waktu presensi
                cursor.execute("""
                    INSERT INTO presensi_mahasiswa (nama_mahasiswa, status, waktu_presensi)
                    VALUES (%s, %s, %s)
                """, (mahasiswa['nama'], 'Alpha', waktu_presensi_alpha))

        conn.commit()

    conn.close()

# Function to initialize and start APScheduler
def start_scheduler():
    scheduler = BackgroundScheduler()
    scheduler.add_job(func=update_alpha_status, trigger="interval", minutes=1)  # Check every 10 minutes
    scheduler.start()

    # Ensure scheduler shuts down when the app exits
    atexit.register(lambda: scheduler.shutdown())

# Function to check for timeout
@app.before_request
def check_timeout():
    if 'user_id' in session:
        now = datetime.now(timezone.utc)
        last_activity = session.get('last_activity', now)

        # Tentukan batas waktu tidak aktif, misalnya 45 menit
        timeout = timedelta(minutes=45)

        # Jika waktu terakhir aktivitas + timeout lebih kecil dari sekarang, logout
        if now - last_activity > timeout:
            session.pop('user_id', None)
            session.pop('role', None)
            session.pop('last_activity', None)
            return redirect(url_for('login'))

        # Perbarui waktu aktivitas terakhir
        session['last_activity'] = now

    # Initialize scheduler if not already started
    if not hasattr(app, 'scheduler_started'):
        start_scheduler()
        app.scheduler_started = True

@app.route('/', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()

        if user and user['password'] == password:
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['last_activity'] = datetime.now()

            if user['role'] == 'mahasiswa':
                return redirect(url_for('home_mahasiswa', login_success=True))
            elif user['role'] == 'mentor':
                return redirect(url_for('home_mentor', login_success=True))
            elif user['role'] == 'dosen':
                return redirect(url_for('home_dosen', login_success=True))
        else:
            error = "Username atau password salah. Silakan coba lagi."

    return render_template('login.html', error=error)

# CREDIT
@app.route('/credit')
def credit():
    return render_template('credit.html')

# USER MAHASISWA
@app.route('/home_mahasiswa')
def home_mahasiswa():
    user_id = session.get('user_id')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT * FROM mahasiswa WHERE id = %s", (user_id,))
    mahasiswa = cursor.fetchone()

    return render_template('home_mahasiswa.html', mahasiswa=mahasiswa)

@app.route('/profil_mahasiswa', methods=['GET', 'POST'])
def profil_mahasiswa():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = None
    try:
        user_id = session['user_id']
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM mahasiswa WHERE id = %s", (user_id,))
        mahasiswa = cursor.fetchone()

        if request.method == 'POST':
            nama = request.form['nama']
            nim = request.form['nim']
            program_studi = request.form['program_studi']
            nama_perguruan_tinggi = request.form['nama_perguruan_tinggi']
            nama_perusahaan = request.form['nama_perusahaan']
            divisi = request.form['divisi']

            if 'foto_profil' in request.files:
                foto = request.files['foto_profil']
                if foto.filename != '':
                    foto_path = os.path.join('static/uploads', foto.filename)
                    foto.save(foto_path)
                    foto_profil = f'uploads/{foto.filename}'
                else:
                    # Set default profile picture if none provided
                    foto_profil = mahasiswa['foto_profil'] if mahasiswa['foto_profil'] else 'uploads/default_user_icon.png'
            else:
                foto_profil = mahasiswa['foto_profil'] if mahasiswa['foto_profil'] else 'uploads/default_user_icon.png'

            cursor.execute("""
                UPDATE mahasiswa 
                SET nama = %s, nim = %s, program_studi = %s, nama_perguruan_tinggi = %s, 
                    nama_perusahaan = %s, divisi = %s, foto_profil = %s 
                WHERE id = %s
            """, (nama, nim, program_studi, nama_perguruan_tinggi, nama_perusahaan, divisi, foto_profil, user_id))
            conn.commit()

            mahasiswa.update({
                'nama': nama,
                'nim': nim,
                'program_studi': program_studi,
                'nama_perguruan_tinggi': nama_perguruan_tinggi,
                'nama_perusahaan': nama_perusahaan,
                'divisi': divisi,
                'foto_profil': foto_profil
            })
    finally:
        if conn is not None:
            conn.close()

    return render_template('profil_mahasiswa.html', mahasiswa=mahasiswa)

@app.route('/scan_qr')
def scan_qr():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Catat waktu presensi saat ini
    waktu_presensi = datetime.now().strftime('%Y-%m-%d')

    # Dapatkan data mahasiswa
    cursor.execute("SELECT nama FROM mahasiswa WHERE id = %s", (user_id,))
    mahasiswa = cursor.fetchone()

    # Periksa apakah mahasiswa sudah melakukan presensi pada tanggal ini
    cursor.execute("""
        SELECT COUNT(*) as total 
        FROM presensi_mahasiswa 
        WHERE nama_mahasiswa = %s AND DATE(waktu_presensi) = %s
    """, (mahasiswa['nama'], waktu_presensi))
    
    presensi_count = cursor.fetchone()['total']

    if presensi_count == 0:
        # Masukkan data presensi ke database jika belum ada presensi pada hari yang sama
        cursor.execute("""
            INSERT INTO presensi_mahasiswa (waktu_presensi, nama_mahasiswa, status)
            VALUES (%s, %s, %s)
        """, (datetime.now().strftime('%Y-%m-%d %H:%M'), mahasiswa['nama'], "Hadir"))
        conn.commit()

    conn.close()

    return redirect(url_for('presensi_mahasiswa'))

@app.route('/presensi_mahasiswa')
def presensi_mahasiswa():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Ambil data mahasiswa berdasarkan user_id dari session
    user_id = session['user_id']
    cursor.execute("SELECT * FROM mahasiswa WHERE id = %s", (user_id,))
    mahasiswa = cursor.fetchone()

    # Ambil semua data presensi dari tabel presensi_mahasiswa
    if mahasiswa:
        # Ambil semua data presensi dari tabel presensi_mahasiswa untuk mahasiswa yang sedang login
        cursor.execute("SELECT * FROM presensi_mahasiswa WHERE nama_mahasiswa = %s ORDER BY waktu_presensi DESC", (mahasiswa['nama'],))
        presensi = cursor.fetchall()
    else:
        presensi = []

    conn.close()

    return render_template('presensi_mahasiswa.html', presensi=presensi, mahasiswa=mahasiswa)

@app.route('/upload_image', methods=['POST'])
def upload_image():
    if 'upload' in request.files:
        file = request.files['upload']
        # Generate filename unik
        filename = None  # Inisialisasi filename sebagai None
        if file and file.filename != '':
            filename = secure_filename(file.filename)
            # Tentukan path penyimpanan
            upload_folder = 'static/uploads/image-ckeditor/'
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)  # Buat direktori jika belum ada
            filepath = os.path.join(upload_folder, filename)
            file.save(filepath)
        
        url = url_for('static', filename='uploads/image-ckeditor/' + filename)
        
        return jsonify({
            "uploaded": True,
            "url": url
        })
    else:
        return jsonify({
            "uploaded": False,
            "error": {
                "message": "Tidak ada file yang di-upload"
            }
        })

@app.route('/update_uraian_tugas/<int:id>', methods=['POST'])
def update_uraian_tugas(id):
    data = request.json
    uraian_tugas = data.get('uraian_tugas', '')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE presensi_mahasiswa 
        SET uraian_tugas = %s 
        WHERE id = %s
    """, (uraian_tugas, id))
    conn.commit()
    conn.close()

    return jsonify({"status": "success"})

@app.route('/permission_mahasiswa')
def permission_mahasiswa():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Ambil data mahasiswa berdasarkan user_id dari session
    user_id = session['user_id']
    cursor.execute("SELECT * FROM mahasiswa WHERE id = %s", (user_id,))
    mahasiswa = cursor.fetchone()

    # Ambil data dari tabel permission
    cursor.execute("SELECT * FROM permission WHERE mahasiswa_id = %s ORDER BY id DESC", (user_id,))
    permissions = cursor.fetchall()

    # Capitalize status for all permissions
    for permission in permissions:
        permission['status'] = permission['status'].capitalize()

    cursor.close()
    conn.close()

    # Pastikan 'mahasiswa' dikirim ke template
    return render_template('permission_mahasiswa.html', mahasiswa=mahasiswa, permissions=permissions)

@app.route('/submit_permission', methods=['POST'])
def submit_permission():
    if 'user_id' not in session:
        return jsonify({"status": "error", "message": "Unauthorized"}), 403

    user_id = session['user_id']  # Ambil user_id dari session (ini mahasiswa_id)

    # Cek apakah mahasiswa sudah mengajukan izin atau sakit pada hari ini
    today = datetime.now().strftime('%Y-%m-%d')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Query untuk mengecek apakah sudah ada permission untuk mahasiswa ini di tanggal hari ini
    cursor.execute("""
        SELECT COUNT(*) AS total FROM permission 
        WHERE mahasiswa_id = %s AND DATE(created_at) = %s
    """, (user_id, today))

    existing_permission = cursor.fetchone()['total']

    if existing_permission > 0:
        # Jika sudah ada permission pada hari ini, kembalikan pesan error
        return jsonify({"status": "error", "message": "Permission already submitted for today."}), 400

    # Jika tidak ada permission pada hari ini, lanjutkan dengan menyimpan data permission baru
    mulai = request.form['mulai']
    selesai = request.form['selesai']
    status = request.form['status'].capitalize()  # pastikan status diubah jadi kapital
    keterangan = request.form['keterangan']
    
    # Dapatkan file yang diupload
    file = request.files['bukti']
    filename = None  # Inisialisasi filename sebagai None
    if file and file.filename != '':
        filename = secure_filename(file.filename)
        # Tentukan path penyimpanan
        upload_folder = 'static/uploads/proof/'
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)  # Buat direktori jika belum ada
        filepath = os.path.join(upload_folder, filename)
        file.save(filepath)
    
    try:
        # Simpan informasi ke database termasuk user_id mahasiswa
        cursor.execute("""
            INSERT INTO permission (mahasiswa_id, tanggal_mulai, tanggal_selesai, status, bukti, keterangan, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW())
        """, (user_id, mulai, selesai, status, filename, keterangan))
        conn.commit()

        # Setelah permission tersimpan, kita tambahkan data ke tabel presensi_mahasiswa
        cursor.execute("SELECT nama FROM mahasiswa WHERE id = %s", (user_id,))
        mahasiswa = cursor.fetchone()
        
        waktu_presensi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            INSERT INTO presensi_mahasiswa (waktu_presensi, nama_mahasiswa, status)
            VALUES (%s, %s, %s)
        """, (waktu_presensi, mahasiswa['nama'], status))
        conn.commit()

        # Kembalikan response JSON dengan data yang akan ditambahkan ke tabel di frontend
        return jsonify({
            "status": "success",
            "permission": {
                "nama": mahasiswa['nama'],
                "mulai": mulai,
                "selesai": selesai,
                "status": status,
                "bukti": filename,
                "keterangan": keterangan
            }
        }), 200

    except Exception as e:
        conn.rollback()  # Jika ada error, rollback perubahan
        return jsonify({"status": "error", "message": "Failed to save permission."}), 500

    finally:
        cursor.close()
        conn.close()


@app.route('/project_mahasiswa', methods=['GET', 'POST'])
def project_mahasiswa():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Ambil data mahasiswa berdasarkan user_id dari session
    user_id = session['user_id']
    cursor.execute("SELECT * FROM mahasiswa WHERE id = %s", (user_id,))
    mahasiswa = cursor.fetchone()

    # Ambil data proyek yang melibatkan mahasiswa ini
    cursor.execute("""
        SELECT * FROM projects WHERE FIND_IN_SET(%s, mahasiswa_ids) ORDER BY id DESC
    """, (str(user_id),))
    projects = cursor.fetchall()

    conn.close()

    # Hitung jumlah proyek ongoing dan selesai
    total_ongoing = sum(1 for project in projects if project['status'] == 0)
    total_completed = sum(1 for project in projects if project['status'] == 1)

    return render_template('project_mahasiswa.html', mahasiswa=mahasiswa, projects=projects, total_ongoing=total_ongoing, total_completed=total_completed)

@app.route('/update_project_status/<int:project_id>', methods=['POST'])
def update_project_status(project_id):
    data = request.json
    new_status = data.get('status')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)  # Ubah ini ke dictionary=True

    # Update project status in the database
    cursor.execute("""
        UPDATE projects 
        SET status = %s 
        WHERE id = %s
    """, (new_status, project_id))
    conn.commit()

    # Fetch updated counts of ongoing and completed projects
    user_id = session.get('user_id')
    cursor.execute("SELECT * FROM projects WHERE FIND_IN_SET(%s, mahasiswa_ids)", (str(user_id),))
    projects = cursor.fetchall()

    total_ongoing = sum(1 for project in projects if project['status'] == 0)
    total_completed = sum(1 for project in projects if project['status'] == 1)

    conn.close()

    # Return updated counts
    return jsonify({"status": "success", "total_ongoing": total_ongoing, "total_completed": total_completed})

# Tambahkan fungsi ini untuk membersihkan HTML dan menampilkan gambar dengan benar
def clean_html_but_allow_images(raw_html):
    # Hapus semua tag kecuali <img> dan <figure>
    cleanr = re.compile(r'(?!(<img|<figure)[^>]*?>)<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return cleantext

# USER MENTOR
@app.route('/home_mentor')
def home_mentor():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    user_id = session['user_id']

    # Ambil data mentor berdasarkan user_id
    cursor.execute("SELECT * FROM mentor WHERE id = %s", (user_id,))
    mentor = cursor.fetchone()

    # Hitung total mahasiswa yang dibimbing oleh mentor
    cursor.execute("SELECT COUNT(*) AS total_mahasiswa FROM mahasiswa")
    total_mahasiswa = cursor.fetchone()['total_mahasiswa']

    # Hitung total tugas on going
    cursor.execute("""
        SELECT COUNT(DISTINCT judul_project) AS total_ongoing
        FROM projects
        WHERE status = 0
    """)
    total_ongoing = cursor.fetchone()['total_ongoing']

    # Hitung total tugas selesai
    cursor.execute("""
        SELECT COUNT(DISTINCT judul_project) AS total_completed
        FROM projects
        WHERE status = 1
    """)
    total_completed = cursor.fetchone()['total_completed']

    # Ambil data presensi mahasiswa untuk ditampilkan di home mentor
    cursor.execute("""
        SELECT id, waktu_presensi, nama_mahasiswa, uraian_tugas, approval_mentor, status
        FROM presensi_mahasiswa
        ORDER BY waktu_presensi DESC
    """)
    presensi = cursor.fetchall()

    # Ganti 'None' dengan 'Belum mengisi task description' pada 'uraian_tugas'
    for record in presensi:
        if not record['uraian_tugas']:  # Cek jika uraian_tugas adalah None atau ''
            record['uraian_tugas'] = 'Belum mengisi task description'

    conn.close()

    return render_template(
        'home_mentor.html',
        mentor=mentor, 
        total_mahasiswa=total_mahasiswa,
        total_ongoing=total_ongoing,
        total_completed=total_completed,
        presensi=presensi
    )

@app.route('/export_excel/<data_type>')
def export_excel(data_type):
    user_id = session.get('user_id')

    if not user_id:
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Pastikan data mahasiswa diambil berdasarkan user_id
    cursor.execute("SELECT nama FROM mahasiswa WHERE id = %s", (user_id,))
    mahasiswa = cursor.fetchone()

    # Logika untuk data presensi
    if data_type == 'presensi':
        # Ambil data presensi hanya untuk mahasiswa yang sedang login
        cursor.execute("SELECT waktu_presensi, nama_mahasiswa, uraian_tugas, approval_mentor FROM presensi_mahasiswa WHERE nama_mahasiswa = %s ORDER BY waktu_presensi ASC", (mahasiswa['nama'],))
        data = cursor.fetchall()

        # Buat file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Presensi Mahasiswa"

        # Tambahkan header
        ws.append(["No", "Waktu Presensi", "Nama Mahasiswa", "Uraian Tugas", "Approval Mentor"])

        # Tambahkan data
        for index, pres in enumerate(data, start=1):
            clean_uraian_tugas = re.sub('<[^<]+?>', '', str(pres['uraian_tugas']))
            ws.append([
                index,  # Nomor urut
                pres['waktu_presensi'].strftime('%d %B %Y %H:%M WIB'),  # Waktu Presensi
                pres['nama_mahasiswa'],  # Nama Mahasiswa
                clean_uraian_tugas,  # Uraian Tugas
                'Approved' if pres['approval_mentor'] else 'Not Approved'  # Approval Mentor
            ])

        # Simpan file Excel
        file_path = 'static/exports/presensi_mahasiswa.xlsx'
        wb.save(file_path)
    
    # Logika untuk data project di project_mahasiswa
    elif data_type == 'project':
        # Ambil data project hanya untuk mahasiswa yang sedang login
        cursor.execute("SELECT judul_project, status, deadline, mahasiswa_ids FROM projects WHERE FIND_IN_SET(%s, mahasiswa_ids)", (str(user_id),))
        data = cursor.fetchall()

        # Buat file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Mahasiswa"

        # Tambahkan header
        ws.append(["No", "Judul Project", "Status", "Deadline", "Mahasiswa"])

        for index, project in enumerate(data, start=1):
            # Convert mahasiswa_ids to names
            cursor.execute("SELECT nama FROM mahasiswa WHERE id IN (%s)" % ','.join(['%s']*len(project['mahasiswa_ids'].split(','))), tuple(project['mahasiswa_ids'].split(',')))
            mahasiswa_names = [mahasiswa['nama'] for mahasiswa in cursor.fetchall()]

            ws.append([
                index,  # Nomor urut
                project['judul_project'],  # Judul Project
                'Completed' if project['status'] == 1 else 'On Going',  # Status
                project['deadline'].strftime('%d %B %Y'),  # Deadline
                ', '.join(mahasiswa_names)  # Mahasiswa
            ])

        # Simpan file Excel
        file_path = 'static/exports/project_mahasiswa.xlsx'
        wb.save(file_path)

    elif data_type == 'project_home_dosen':
        # Logika untuk data project di home_dosen
        cursor.execute("SELECT judul_project, status, deadline, mahasiswa_ids FROM projects ORDER BY id ASC")
        data = cursor.fetchall()

        # Buat file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Mahasiswa"

        # Tambahkan header
        ws.append(["No", "Judul Project", "Status", "Deadline", "Nama Mahasiswa"])

        for index, project in enumerate(data, start=1):
            # Mengonversi mahasiswa_ids menjadi nama mahasiswa
            mahasiswa_ids = project['mahasiswa_ids'].split(',')
            if mahasiswa_ids:
                cursor.execute("SELECT nama FROM mahasiswa WHERE id IN (%s)" % ','.join(['%s']*len(mahasiswa_ids)), tuple(mahasiswa_ids))
                mahasiswa_names = [mahasiswa['nama'] for mahasiswa in cursor.fetchall()]
                mahasiswa_names_str = ', '.join(mahasiswa_names)
            else:
                mahasiswa_names_str = ''

            # Menambahkan baris ke worksheet Excel
            ws.append([
                index,  # Nomor urut
                project['judul_project'],  # Judul Project
                'Completed' if project['status'] == 1 else 'On Going',  # Status
                project['deadline'].strftime('%d %B %Y'),  # Deadline
                mahasiswa_names_str  # Nama Mahasiswa (dalam format string)
            ])

        # Simpan file Excel
        file_path = 'static/exports/data_project_mahasiswa.xlsx'
        wb.save(file_path)

        return send_file(file_path, as_attachment=True, download_name='data_project_mahasiswa.xlsx')


    # Logika untuk data mentor (contoh yang sudah ada)
    elif data_type == 'mentor':
        # Ambil semua data presensi (tanpa filter untuk mentor)
        cursor.execute("SELECT waktu_presensi, nama_mahasiswa, uraian_tugas, approval_mentor FROM presensi_mahasiswa ORDER BY waktu_presensi ASC")
        data = cursor.fetchall()

        # Buat file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Presensi Mahasiswa"

        # Tambahkan header
        ws.append(["No", "Waktu Presensi", "Nama Mahasiswa", "Uraian Tugas", "Approval Mentor"])

        for index, pres in enumerate(data, start=1):
            clean_uraian_tugas = re.sub('<[^<]+?>', '', str(pres['uraian_tugas']))
            ws.append([
                index,  # Nomor urut
                pres['waktu_presensi'].strftime('%d %B %Y %H:%M WIB'),  # Waktu Presensi
                pres['nama_mahasiswa'],  # Nama Mahasiswa
                clean_uraian_tugas,  # Uraian Tugas
                'Approved' if pres['approval_mentor'] else 'Not Approved'  # Approval Mentor
            ])

        # Simpan file Excel
        file_path = 'static/exports/data_presensi_mahasiswa.xlsx'
        wb.save(file_path)

    else:
        return "Invalid data type", 400

    # Kirim file ke user
    return send_file(file_path, as_attachment=True, download_name=file_path.split('/')[-1])


@app.route('/update_approval_mentor/<int:presensi_id>', methods=['POST'])
def update_approval_mentor(presensi_id):
    data = request.json
    approval_mentor = data.get('approval_mentor', 0)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE presensi_mahasiswa 
        SET approval_mentor = %s 
        WHERE id = %s
    """, (approval_mentor, presensi_id))
    conn.commit()
    conn.close()

    return jsonify({"status": "success"})

@app.route('/profil_mentor', methods=['GET', 'POST'])
def profil_mentor():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = None
    try:
        user_id = session['user_id']
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM mentor WHERE id = %s", (user_id,))
        mentor = cursor.fetchone()

        if request.method == 'POST':
            nama = request.form['nama']
            email = request.form['email']
            nama_perusahaan = request.form['nama_perusahaan']
            divisi = request.form['divisi']

            if 'foto_profil' in request.files:
                foto = request.files['foto_profil']
                if foto.filename != '':
                    foto_path = os.path.join('static/uploads', foto.filename)
                    foto.save(foto_path)
                    foto_profil = f'uploads/{foto.filename}'
                else:
                    foto_profil = mentor['foto_profil']
            else:
                foto_profil = mentor['foto_profil']

            cursor.execute("""
                UPDATE mentor 
                SET nama = %s, email = %s, nama_perusahaan = %s, divisi = %s, foto_profil = %s
                WHERE id = %s
            """, (nama, email, nama_perusahaan, divisi, foto_profil, user_id))
            conn.commit()

            mentor.update({
                'nama': nama,
                'email': email,
                'nama_perusahaan': nama_perusahaan,
                'divisi': divisi,
                'foto_profil': foto_profil
            })
    finally:
        if conn is not None:
            conn.close()

    return render_template('profil_mentor.html', mentor=mentor)

@app.route('/project_mentor')
def project_mentor():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    user_id = session['user_id']

    # Ambil data mentor berdasarkan user_id
    cursor.execute("SELECT * FROM mentor WHERE id = %s", (user_id,))
    mentor = cursor.fetchone()

    # Hitung total mahasiswa yang dibimbing oleh mentor
    cursor.execute("SELECT COUNT(*) AS total_mahasiswa FROM mahasiswa")
    total_mahasiswa = cursor.fetchone()['total_mahasiswa']

    # Hitung total tugas on going
    cursor.execute("""
        SELECT COUNT(DISTINCT judul_project) AS total_ongoing
        FROM projects
        WHERE status = 0
    """)
    total_ongoing = cursor.fetchone()['total_ongoing']

    # Hitung total tugas selesai
    cursor.execute("""
        SELECT COUNT(DISTINCT judul_project) AS total_completed
        FROM projects
        WHERE status = 1
    """)
    total_completed = cursor.fetchone()['total_completed']

    # Ambil data project mahasiswa
    cursor.execute("SELECT p.id, p.judul_project, p.status, p.deadline, p.mahasiswa_ids FROM projects p ORDER BY id DESC")
    projects = cursor.fetchall()

    # Ambil data mahasiswa terkait
    for project in projects:
        mahasiswa_ids = project['mahasiswa_ids']
        if mahasiswa_ids:  # Pastikan mahasiswa_ids bukan None
            mahasiswa_ids_list = mahasiswa_ids.split(',')  # Pisahkan string menjadi list ID
            cursor.execute("SELECT nama FROM mahasiswa WHERE id IN (%s)" % ','.join(['%s']*len(mahasiswa_ids_list)), tuple(mahasiswa_ids_list))
            mahasiswa_names = cursor.fetchall()
            project['mahasiswa'] = ', '.join([m['nama'] for m in mahasiswa_names])
        else:
            project['mahasiswa'] = "Tidak ada mahasiswa"

    # Ambil data semua mahasiswa untuk dropdown
    cursor.execute("SELECT id, nama FROM mahasiswa")
    mahasiswa_list = cursor.fetchall()

    conn.close()

    return render_template(
        'project_mentor.html',
        mentor=mentor,
        total_mahasiswa=total_mahasiswa,
        total_ongoing=total_ongoing,
        total_completed=total_completed,
        projects=projects,
        mahasiswa_list=mahasiswa_list)

@app.route('/tambah_project', methods=['POST'])
def tambah_project():
    data = request.form
    judul_project = data['judul_project']
    deadline = data['deadline']
    mahasiswa_ids = request.form.getlist('mahasiswa')  # Ambil semua checkbox yang dipilih
    
    # Gabungkan daftar ID mahasiswa menjadi satu string dengan koma sebagai pemisah
    mahasiswa_ids_str = ",".join(mahasiswa_ids)

    conn = get_db_connection()
    cursor = conn.cursor()

    # Insert project ke tabel projects
    cursor.execute("""
        INSERT INTO projects (judul_project, status, deadline, mahasiswa_ids)
        VALUES (%s, %s, %s, %s)
    """, (judul_project, 0, deadline, mahasiswa_ids_str))

    conn.commit()
    conn.close()

    return redirect(url_for('project_mentor'))

@app.route('/edit_project_data/<int:id>', methods=['GET'])
def edit_project_data(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Ambil data project berdasarkan ID
    cursor.execute("SELECT * FROM projects WHERE id = %s", (id,))
    project = cursor.fetchone()

    # Format tanggal deadline menjadi YYYY-MM-DD
    if project['deadline']:
        project['deadline'] = project['deadline'].strftime('%Y-%m-%d')

    # Pastikan mahasiswa_ids tidak None
    if project['mahasiswa_ids']:
        project['mahasiswa_ids'] = project['mahasiswa_ids'].split(',')
    else:
        project['mahasiswa_ids'] = []

    conn.close()

    return jsonify(project)

@app.route('/update_project/<int:id>', methods=['POST'])
def update_project(id):
    data = request.form
    judul_project = data['judul_project']
    deadline = data['deadline']
    mahasiswa_ids = ','.join(data.getlist('mahasiswa'))  # Menggabungkan ID mahasiswa menjadi string

    conn = get_db_connection()
    cursor = conn.cursor()

    # Update project ke tabel projects
    cursor.execute("""
        UPDATE projects
        SET judul_project = %s, deadline = %s, mahasiswa_ids = %s
        WHERE id = %s
    """, (judul_project, deadline, mahasiswa_ids, id))

    conn.commit()
    conn.close()

    return redirect(url_for('project_mentor'))

@app.route('/delete_project/<int:id>', methods=['POST'])
def delete_project(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Delete the project
    cursor.execute("DELETE FROM projects WHERE id = %s", (id,))

    conn.commit()
    conn.close()

    return redirect(url_for('project_mentor'))

# USER DOSEN
@app.route('/home_dosen')
def home_dosen():
    if 'user_id' not in session or session.get('role') != 'dosen':
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    user_id = session['user_id']

    # Ambil data dosen berdasarkan user_id
    cursor.execute("SELECT * FROM dosen WHERE id = %s", (user_id,))
    dosen = cursor.fetchone()

    # Ambil total mahasiswa
    cursor.execute("SELECT COUNT(*) AS total_mahasiswa FROM mahasiswa")
    total_mahasiswa = cursor.fetchone()['total_mahasiswa']

    # Ambil total mentor
    cursor.execute("SELECT COUNT(*) AS total_mentor FROM mentor")
    total_mentor = cursor.fetchone()['total_mentor']

    # Ambil total project on going
    cursor.execute("SELECT COUNT(DISTINCT judul_project) AS total_ongoing FROM projects WHERE status = 0")
    total_ongoing = cursor.fetchone()['total_ongoing']

    # Ambil total project completed
    cursor.execute("SELECT COUNT(DISTINCT judul_project) AS total_completed FROM projects WHERE status = 1")
    total_completed = cursor.fetchone()['total_completed']

    conn.close()

    return render_template('home_dosen.html', dosen=dosen, total_mahasiswa=total_mahasiswa, total_mentor=total_mentor, 
                           total_ongoing=total_ongoing, total_completed=total_completed)

@app.route('/kehadiran_hari_ini', methods=['GET'])
def kehadiran_hari_ini():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get today's date in the correct format (YYYY-MM-DD)
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Query untuk menghitung jumlah mahasiswa berdasarkan status untuk hari ini
    cursor.execute("""
        SELECT status, COUNT(*) as count 
        FROM presensi_mahasiswa 
        WHERE DATE(waktu_presensi) = %s 
        GROUP BY status
    """, (today,))
    
    presensi_hari_ini = cursor.fetchall()
    
    # Buat dictionary untuk menyimpan jumlah status
    result = {
        'Hadir': 0,
        'Sakit': 0,
        'Izin': 0,
        'Alpha': 0,
    }

    # Isi result dictionary dengan jumlah status dari query
    for row in presensi_hari_ini:
        result[row['status']] = row['count']

    conn.close()
    
    # Return hasilnya sebagai JSON untuk frontend
    return jsonify(result)


@app.route('/get_project_status_data', methods=['GET'])
def get_project_status_data():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch the project statuses
    cursor.execute("SELECT status FROM projects")
    projects = cursor.fetchall()
    conn.close()

    # Calculate the total number of ongoing and completed projects
    total_ongoing = sum(1 for project in projects if project['status'] == 0)
    total_completed = sum(1 for project in projects if project['status'] == 1)

    # Return the counts as JSON
    result = {
        'total_ongoing': total_ongoing,
        'total_completed': total_completed
    }
    return jsonify(result)

@app.route('/detail_presensi')
def detail_presensi():
    if 'user_id' not in session or session.get('role') != 'dosen':
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    user_id = session['user_id']

    # Ambil data dosen berdasarkan user_id
    cursor.execute("SELECT * FROM dosen WHERE id = %s", (user_id,))
    dosen = cursor.fetchone()

    # Ambil kata kunci pencarian dari form pencarian
    search_query = request.args.get('search_query', '')

    # Ambil data presensi sesuai pencarian
    if search_query:
        cursor.execute("""
            SELECT * FROM presensi_mahasiswa 
            WHERE nama_mahasiswa LIKE %s 
            ORDER BY waktu_presensi DESC
        """, ('%' + search_query + '%',))
        presensi_mahasiswa = cursor.fetchall()
    else:
        cursor.execute("SELECT * FROM presensi_mahasiswa ORDER BY waktu_presensi DESC")
        presensi_mahasiswa = cursor.fetchall()

    # Ganti 'None' atau string kosong dengan 'Belum mengisi task description'
    for presensi in presensi_mahasiswa:
        if not presensi['uraian_tugas']:  # If uraian_tugas is empty or None
            presensi['uraian_tugas'] = 'Belum mengisi task description'

    conn.close()

    return render_template('detail_presensi.html', dosen=dosen, presensi_mahasiswa=presensi_mahasiswa, search_query=search_query)

@app.route('/detail_project')
def detail_project():
    if 'user_id' not in session or session.get('role') != 'dosen':
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    user_id = session['user_id']

    # Ambil data dosen berdasarkan user_id
    cursor.execute("SELECT * FROM dosen WHERE id = %s", (user_id,))
    dosen = cursor.fetchone()

    # Ambil kata kunci pencarian dari form pencarian
    search_query = request.args.get('search_query', '')

    # Ambil data proyek sesuai pencarian
    if search_query:
        cursor.execute("""
            SELECT p.*, GROUP_CONCAT(m.nama SEPARATOR ', ') AS mahasiswa_names 
            FROM projects p
            LEFT JOIN mahasiswa m ON FIND_IN_SET(m.id, p.mahasiswa_ids)
            WHERE p.judul_project LIKE %s
            GROUP BY p.id
            ORDER BY id DESC
            """, ('%' + search_query + '%',))
        projects = cursor.fetchall()
    else:
        cursor.execute("""
            SELECT p.*,
            IFNULL(GROUP_CONCAT(m.nama SEPARATOR ', '), 'Tidak ada mahasiswa') AS mahasiswa_names
            FROM projects p
            LEFT JOIN mahasiswa m ON FIND_IN_SET(m.id, p.mahasiswa_ids)
            GROUP BY p.id
            ORDER BY id DESC
            """)
        projects = cursor.fetchall()
        

    conn.close()

    return render_template('detail_project.html', dosen=dosen, projects=projects, search_query=search_query)

@app.route('/profil_dosen', methods=['GET', 'POST'])
def profil_dosen():
    if 'user_id' not in session or session.get('role') != 'dosen':
        return redirect(url_for('login'))

    conn = None
    try:
        user_id = session['user_id']
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM dosen WHERE id = %s", (user_id,))
        dosen = cursor.fetchone()

        if request.method == 'POST':
            nama = request.form['nama']
            nidn = request.form['nidn']
            email = request.form['email']
            jabatan = request.form['jabatan']

            if 'foto_profil' in request.files:
                foto = request.files['foto_profil']
                if foto.filename != '':
                    foto_path = os.path.join('static/uploads', foto.filename)
                    foto.save(foto_path)
                    foto_profil = f'uploads/{foto.filename}'
                else:
                    foto_profil = dosen['foto_profil']
            else:
                foto_profil = dosen['foto_profil']

            cursor.execute("""
                UPDATE dosen 
                SET nama = %s, nidn = %s, email = %s, jabatan = %s, foto_profil = %s
                WHERE id = %s
            """, (nama, nidn, email, jabatan, foto_profil, user_id))
            conn.commit()

            dosen.update({
                'nama': nama,
                'nidn': nidn,
                'email': email,
                'jabatan': jabatan,
                'foto_profil': foto_profil
            })
    finally:
        if conn is not None:
            conn.close()

    return render_template('profil_dosen.html', dosen=dosen)

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('role', None)
    session.pop('last_activity', None)
    return redirect(url_for('login', logout_success=True))

if __name__ == '__main__':
    app.run(debug=True)
    app.run(debug=True)
